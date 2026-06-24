#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import re
import sys
import time
import uuid
import argparse
import datetime
from typing import List, Dict, Optional, Any
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock

import netmiko
import openpyxl
import encodings.idna  # noqa: F401  (线程内编码问题预加载)
from tqdm import tqdm
from netmiko import (
    NetmikoTimeoutException,
    NetmikoAuthenticationException,
)
from netmiko.ssh_dispatcher import CLASS_MAPPER

# ---------------------------------------------------------------------------
# 环境配置
# ---------------------------------------------------------------------------
os.environ["NO_COLOR"] = "1"
write_lock = Lock()
DEFAULT_THREADS = min(900, max(4, (os.cpu_count() or 4)))

SUPPORTED_DEVICE_TYPES = set(CLASS_MAPPER.keys())

# 需要手写分页的 device_type（Netmiko send_command 不会自动应答 --More--）
MANUAL_PAGER_TYPES = {'generic_termserver', 'terminal_server', 'generic'}

# ---------------------------------------------------------------------------
# 全局正则（大小写不敏感 flag 统一放 re.compile 参数，避免 (?i) 中置报错）
# ---------------------------------------------------------------------------
PAGER_RE = re.compile(
    '|'.join([
        r'--?\s*more\s*--?',           # --More--, - More -
        r'more\s*[:?]',                # More:, More?
        r'\(space.*?to\s+continue\)',  # (space to continue)
        r'press\s+any\s+key',
        r'press\s+<?space>?',
        r'hit\s+any\s+key',
    ]),
    re.IGNORECASE,
)
ANSI_RE = re.compile(r'\x1b\[[0-9;?]*[A-Za-z]')   # CSI 序列：\x1b[7m \x1b[K 等
ANSI_OTHER_RE = re.compile(r'\x1b[()][AB0]')      # 字符集切换
BACKSPACE_RE = re.compile(r'.\x08')               # 退格覆盖：前一字符 + \x08

# ---------------------------------------------------------------------------
# 设备类型别名映射
# ---------------------------------------------------------------------------
DEVICE_TYPE_ALIASES = {
    'cisco': 'cisco_ios', 'cisco_switch': 'cisco_ios', 'cisco_router': 'cisco_ios',
    'nexus': 'cisco_nxos', 'asa': 'cisco_asa', 'ios_xe': 'cisco_xe', 'ios_xr': 'cisco_xr',
    'huawei': 'huawei', 'vrp': 'huawei', 'vrpv8': 'huawei_vrpv8',
    'hp': 'hp_comware', 'comware': 'hp_comware', 'h3c': 'hp_comware',
    'procurve': 'hp_procurve', 'aruba': 'aruba_os',
    'juniper': 'juniper', 'junos': 'juniper', 'srx': 'juniper_screenos',
    'fortinet': 'fortinet', 'fortigate': 'fortinet', 'fortios': 'fortinet',
    'paloalto': 'paloalto_panos', 'panos': 'paloalto_panos',
    'dell': 'dell_force10', 'force10': 'dell_force10',
    'mikrotik': 'mikrotik_routeros', 'routeros': 'mikrotik_routeros',
    'nokia': 'nokia_sros', 'sros': 'nokia_sros', 'f5': 'f5_tmsh', 'bigip': 'f5_tmsh',
    'linux': 'linux', 'ubuntu': 'linux', 'centos': 'linux',
    'generic_termserver': 'generic_termserver',
    'terminal_server': 'generic_termserver',
}


def resolve_device_type(raw: Any) -> str:
    """把别名或原始字符串解析为 Netmiko 标准 device_type。"""
    if raw is None:
        return 'generic_termserver'
    key = str(raw).strip().lower()
    if not key:
        return 'generic_termserver'
    if key in SUPPORTED_DEVICE_TYPES:
        return key
    if key in DEVICE_TYPE_ALIASES:
        return DEVICE_TYPE_ALIASES[key]
    return 'generic_termserver'


# ---------------------------------------------------------------------------
# 厂商连接参数（按解析后的 device_type 前缀匹配）
# ---------------------------------------------------------------------------
VENDOR_TIMEOUTS = {
    'cisco':   {'timeout': 25, 'banner_timeout': 15, 'auth_timeout': 10},
    'huawei':  {'timeout': 30, 'banner_timeout': 15, 'auth_timeout': 12},
    'hp':      {'timeout': 30, 'banner_timeout': 15, 'auth_timeout': 12},
    'juniper': {'timeout': 30, 'banner_timeout': 15, 'auth_timeout': 12},
    'default': {'timeout': 25, 'banner_timeout': 15, 'auth_timeout': 10},
}


def get_conn_extra(device_type: str) -> Dict[str, Any]:
    for prefix, cfg in VENDOR_TIMEOUTS.items():
        if prefix != 'default' and device_type.startswith(prefix):
            return dict(cfg)
    return dict(VENDOR_TIMEOUTS['default'])


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------
def thread_initializer() -> None:
    """线程初始化（解决编码问题）"""
    import encodings.idna
    encodings.idna.__name__  # 防止被优化


def sanitize_filename(name: str) -> str:
    """生成安全文件名"""
    return re.sub(r'[\\/*?:"<>|]', '', str(name)).strip()[:60]


def log_error(host: str, msg: str) -> None:
    """统一错误日志（控制台 + error_log 文件）"""
    line = f"[{datetime.datetime.now():%Y-%m-%d %H:%M:%S}] {host} {msg}"
    print(line)
    with write_lock:
        with open("error_log.txt", "a", encoding="utf-8") as f:
            f.write(line + "\n")


def validate_device_data(device: Dict[str, str], row_idx: int) -> None:
    """验证设备数据完整性"""
    required = ['host', 'username', 'password', 'device_type']
    if missing := [f for f in required if not device.get(f)]:
        raise ValueError(f"Row {row_idx} 缺失字段: {', '.join(missing)}")


# ---------------------------------------------------------------------------
# 输出目录与保存（恢复原本的 result_日期 / IP_主机名.txt）
# ---------------------------------------------------------------------------
def get_output_dir() -> str:
    """输出目录：result_当天日期，例如 result_20260624。"""
    date_str = datetime.datetime.now().strftime('%Y%m%d')
    out_dir = f"result_{date_str}"
    os.makedirs(out_dir, exist_ok=True)
    return out_dir


def save_result(device: Dict[str, str], content: str, out_dir: str) -> None:
    """保存为 IP_主机名.txt。"""
    host = sanitize_filename(device.get('host', 'unknown'))
    hostname = sanitize_filename(device.get('hostname', '') or 'nohost')
    filename = f"{host}_{hostname}.txt"
    path = os.path.join(out_dir, filename)

    with write_lock:
        with open(path, 'w', encoding='utf-8') as f:
            f.write(content)
    print(f"{device['host']} [OK] 已保存 -> {path}")


# ---------------------------------------------------------------------------
# Excel 加载
# ---------------------------------------------------------------------------
def load_excel(excel_file: str, sheet_name: str = 'Sheet1') -> List[Dict[str, str]]:
    """加载Excel设备清单（数值单元格安全转换 + 列校验）"""
    devices: List[Dict[str, str]] = []
    wb = None
    try:
        wb = openpyxl.load_workbook(excel_file, read_only=True)

        if sheet_name not in wb.sheetnames:
            raise ValueError(f"工作表 '{sheet_name}' 不存在")
        sheet = wb[sheet_name]

        headers = [str(cell.value).lower().strip() if cell.value is not None else ""
                   for cell in sheet[1]]
        required = ['host', 'username', 'password', 'device_type']
        if missing := [f for f in required if f not in headers]:
            raise ValueError(f"缺少必要列: {', '.join(missing)}")

        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
            device = {
                headers[i]: ("" if cell is None else str(cell).strip())
                for i, cell in enumerate(row)
                if i < len(headers) and headers[i]
            }
            # 跳过整行空白
            if not any(device.values()):
                continue
            validate_device_data(device, row_idx)
            # 规范化 device_type
            device['device_type'] = resolve_device_type(device.get('device_type'))
            devices.append(device)

        return devices
    except Exception as e:
        print(f"Excel处理失败: {str(e)}")
        sys.exit(1)
    finally:
        if wb:
            wb.close()


# ---------------------------------------------------------------------------
# 设备连接
# ---------------------------------------------------------------------------
def connect_device(device: Dict[str, str]) -> Optional[netmiko.BaseConnection]:
    """设备连接（自动生成独立 debug 日志文件）"""
    dtype = device['device_type']
    params = {
        'device_type': dtype,
        'host': device['host'],
        'username': device['username'],
        'password': device['password'],
        'secret': device.get('secret', ''),
        'fast_cli': False,
        **get_conn_extra(dtype),
    }
    if device.get('port'):
        try:
            params['port'] = int(device['port'])
        except ValueError:
            pass

    # 动态生成设备专属 debug 日志路径
    if device.get('debug'):
        debug_dir = os.path.join("debug_logs", datetime.datetime.now().strftime('%Y%m%d'))
        os.makedirs(debug_dir, exist_ok=True)
        log_file = f"{sanitize_filename(device['host'])}_{uuid.uuid4().hex[:6]}.log"
        params['session_log'] = os.path.join(debug_dir, log_file)

    try:
        conn = netmiko.ConnectHandler(**params)
        if params['secret']:
            conn.enable()
        return conn
    except (NetmikoTimeoutException, NetmikoAuthenticationException) as e:
        log_error(device['host'], f"{e.__class__.__name__}: {str(e)}")
    except Exception as e:
        log_error(device['host'], f"连接异常: {str(e)}")
    return None


# ---------------------------------------------------------------------------
# 终端服务器手写分页
# ---------------------------------------------------------------------------
def send_command_termserver(
    conn: netmiko.BaseConnection,
    cmd: str,
    delay_factor: float = 1.0,
    max_loops: int = 2000,
) -> str:
    """对终端服务器/generic 设备手动应答分页（Netmiko 不会自动处理）。"""
    remote = getattr(conn, "remote_conn", None)

    def drain() -> str:
        buf = []
        while True:
            chunk = conn.read_channel()
            if chunk:
                buf.append(chunk)
                time.sleep(0.08 * delay_factor)
                continue
            if remote is not None and remote.recv_ready():
                time.sleep(0.08 * delay_factor)
                continue
            break
        return ''.join(buf)

    output: List[str] = []
    conn.write_channel(cmd + conn.RETURN)
    time.sleep(0.3 * delay_factor)

    first = drain()
    if first:
        output.append(first)

    seen_pager = False
    empty_reads = 0

    for _ in range(max_loops):
        tail = ANSI_RE.sub('', ''.join(output))[-512:]

        if PAGER_RE.search(tail):
            seen_pager = True
            conn.write_channel(' ')
            time.sleep(0.18 * delay_factor)
            chunk = drain()
            if chunk:
                output.append(chunk)
                empty_reads = 0
                continue
            empty_reads += 1
        else:
            chunk = drain()
            if chunk:
                output.append(chunk)
                empty_reads = 0
                time.sleep(0.08 * delay_factor)
                continue
            empty_reads += 1
            if seen_pager and empty_reads >= 4:
                break
            if not seen_pager and empty_reads >= 6:
                break

    return clean_pager_output(''.join(output))


def clean_pager_output(raw: str) -> str:
    """清理分页回显：去 ANSI、去退格重绘、删 pager 整行、压缩多余空行。"""
    text = raw

    # 1. 退格覆盖：循环处理（连续退格 aaa\x08\x08\x08 单次删不净）
    prev = None
    while prev != text:
        prev = text
        text = BACKSPACE_RE.sub('', text)

    # 2. 去 ANSI 控制序列
    text = ANSI_RE.sub('', text)
    text = ANSI_OTHER_RE.sub('', text)

    # 3. 统一换行，去掉孤立的 \r（设备重绘用）
    text = text.replace('\r\n', '\n').replace('\r', '\n')

    # 4. 按行处理：删整行 pager、清行内夹杂 pager、去行尾空白
    cleaned_lines = []
    for line in text.split('\n'):
        stripped = line.strip()
        if stripped and PAGER_RE.fullmatch(stripped):
            continue
        line = PAGER_RE.sub('', line)
        cleaned_lines.append(line.rstrip())

    text = '\n'.join(cleaned_lines)

    # 5. 压缩 3 行以上连续空行为 1 个空行
    text = re.sub(r'\n{3,}', '\n\n', text)

    return text.strip('\n')


# ---------------------------------------------------------------------------
# 执行命令
# ---------------------------------------------------------------------------
def get_device_commands(device: Dict[str, str], global_cmds: List[str]) -> List[str]:
    """优先用行内命令，行内为空才回退全局 -c。"""
    raw = device.get('mult_command') or device.get('commands') or ''
    cmds = [c.strip() for c in re.split(r'[;\n\r]+', raw) if c.strip()]
    if cmds:
        return cmds
    return list(global_cmds)


def execute_commands(
    device: Dict[str, str],
    global_cmds: List[str],
    out_dir: str,
) -> Optional[str]:
    """执行命令、清理回显并保存 txt。"""
    cmds = get_device_commands(device, global_cmds)
    if not cmds:
        print(f"{device['host']} [WARN] 无有效命令（行内为空且无 -c 兜底）")
        return None

    if not (conn := connect_device(device)):
        return None

    dtype = device['device_type']
    use_manual_pager = dtype in MANUAL_PAGER_TYPES

    try:
        with conn:
            # 获取主机名
            prompt = conn.find_prompt().strip()
            m = re.search(r'([\w.\-]+)\s*[#>$\]]\s*$', prompt)
            device['hostname'] = m.group(1) if m else sanitize_filename(device['host'])

            blocks = []
            for cmd in cmds:
                if use_manual_pager:
                    out = send_command_termserver(conn, cmd)
                else:
                    out = conn.send_command(
                        cmd,
                        read_timeout=int(device.get('readtime', 20) or 20),
                        strip_prompt=False,
                        strip_command=False,
                    )
                    out = clean_pager_output(out)

                blocks.append(
                    f"{'=' * 70}\n"
                    f"# {device['host']} ({device['hostname']}) >> {cmd}\n"
                    f"{'-' * 70}\n"
                    f"{out}\n"
                )

            content = "\n".join(blocks)

        save_result(device, content, out_dir)
        return content

    except Exception as e:
        log_error(device['host'], f"执行异常: {str(e)}")
        return None


# ---------------------------------------------------------------------------
# 主流程
# ---------------------------------------------------------------------------
def main() -> None:
    parser = argparse.ArgumentParser(description="批量网络设备命令执行（Netmiko + Excel）")
    parser.add_argument('-f', '--file', required=True, help="设备清单 Excel 文件")
    parser.add_argument('-s', '--sheet', default='Sheet1', help="工作表名（默认 Sheet1）")
    parser.add_argument('-c', '--command', default='',
                        help="全局兜底命令（多条用分号分隔），仅当某行无命令时使用")
    parser.add_argument('-t', '--threads', type=int, default=DEFAULT_THREADS,
                        help=f"并发线程数（默认 {DEFAULT_THREADS}）")
    args = parser.parse_args()

    global_cmds = [c.strip() for c in re.split(r'[;\n\r]+', args.command) if c.strip()]

    devices = load_excel(args.file, args.sheet)
    if not devices:
        print("无可执行设备，退出。")
        sys.exit(0)

    out_dir = get_output_dir()
    print(f"共 {len(devices)} 台设备，输出目录：{out_dir}")

    ok = 0
    with ThreadPoolExecutor(
        max_workers=min(args.threads, len(devices)),
        initializer=thread_initializer,
    ) as pool:
        futures = {
            pool.submit(execute_commands, dev, global_cmds, out_dir): dev
            for dev in devices
        }
        for fut in tqdm(as_completed(futures), total=len(futures), desc="执行进度"):
            dev = futures[fut]
            try:
                if fut.result():
                    ok += 1
            except Exception as e:
                log_error(dev['host'], f"线程异常: {str(e)}")

    print(f"\n完成：成功 {ok} / 总计 {len(devices)}，结果保存在 {out_dir}")


if __name__ == '__main__':
    main()
